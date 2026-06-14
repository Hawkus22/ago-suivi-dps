# ui_main.py
import sys
import pandas as pd
from PyQt6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QPushButton, QFileDialog, QTableWidget,
                             QTableWidgetItem, QHeaderView, QMessageBox, QMenu,
                             QAbstractItemView, QApplication, QComboBox, QTextBrowser,
                             QDialog, QDialogButtonBox)
from PyQt6.QtCore import Qt, QRect, QSize, QStandardPaths
from PyQt6.QtGui import QColor, QPainter, QFont, QPen
from database import get_conn
from config import ANTENNES_ORDRE, JOURS_SEMAINE, COLOR_ORANGE, COLOR_GREEN, COLOR_RENFORT_OK, COLOR_RENFORT_KO, resource_path
from version import APP_NAME, VERSION, BUILD, AUTHOR
from import_handler import importer_evenements
from renfort_engine import suggerer_renforts, ajouter_renforts, toutes_disponibilites
from ui_popup import RenfortPopup

_DPS_ID_ROLE = Qt.ItemDataRole.UserRole


class TwoLineHeaderView(QHeaderView):
    """En-tête sur 2 lignes : Jour (haut, fusionné) + Sous-colonnes (bas).

    Les colonnes de self.table commencent directement par les jours
    (la colonne Antenne est dans un tableau séparé figé à gauche).
    Disposition : jour_idx * 4 + sous_col_idx (0=DPS 1=Engagés 2=Manque 3=Σ).
    """

    SOUS_COLONNES = ["DPS", "Engagés", "Manque", "Σ"]
    HEADER_HEIGHT = 60

    def __init__(self, parent=None):
        super().__init__(Qt.Orientation.Horizontal, parent)
        self.setFixedHeight(self.HEADER_HEIGHT)
        self.setSectionsClickable(False)
        self.setSectionsMovable(False)
        self.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

    def sizeHint(self):
        return QSize(super().sizeHint().width(), self.HEADER_HEIGHT)

    _TOOLTIPS_SOUS_COL = [
        "Nom du DPS / renfort",
        "Nombre d'IS confirmés sur ce DPS\n(renforts inclus pour les DPS principaux)",
        "Nombre d'IS encore nécessaires pour couvrir ce DPS",
        "Effectif total requis (Σ)",
    ]

    def event(self, e):
        from PyQt6.QtCore import QEvent
        if e.type() == QEvent.Type.ToolTip:
            pos = e.pos()
            idx = self.logicalIndexAt(pos)
            if idx >= 0:
                self.setToolTip(self._TOOLTIPS_SOUS_COL[idx % 4])
        return super().event(e)

    def paintSection(self, painter, rect, logical_index):
        if not rect.isValid():
            return

        FOND  = QColor(30, 60, 114)
        TEXTE = QColor(255, 255, 255)
        BORD  = QColor(100, 120, 150)
        hl    = self.height() // 2

        painter.save()
        painter.fillRect(rect, FOND)
        painter.setPen(TEXTE)

        sous_col = self.SOUS_COLONNES[logical_index % 4]
        painter.setFont(QFont("Arial", 9, QFont.Weight.Bold))
        painter.drawText(
            QRect(rect.x(), rect.y() + hl, rect.width(), hl),
            Qt.AlignmentFlag.AlignCenter, sous_col
        )
        painter.setPen(QPen(BORD, 1))
        painter.drawLine(rect.left(), rect.top() + hl, rect.right(), rect.top() + hl)
        painter.drawLine(rect.right(), rect.top(), rect.right(), rect.bottom())
        painter.restore()

    def paintEvent(self, event):
        super().paintEvent(event)

        painter = QPainter(self.viewport())
        if not painter.isActive():
            return

        TEXTE = QColor(255, 255, 255)
        hl = self.height() // 2

        painter.setPen(TEXTE)
        for jour_idx, jour in enumerate(JOURS_SEMAINE):
            col_start = jour_idx * 4
            if col_start >= self.count():
                break
            x = self.sectionViewportPosition(col_start)
            nb_visible = min(4, self.count() - col_start)
            w = sum(self.sectionSize(col_start + i) for i in range(nb_visible))
            painter.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            painter.drawText(QRect(x, 0, w, hl), Qt.AlignmentFlag.AlignCenter, jour)

        painter.end()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PHARE - Suivi Départemental des DPS")
        self.resize(1400, 800)
        self.current_semaine = None
        self.init_ui()

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        # Menu bar
        menubar = self.menuBar()
        menu_fichier = menubar.addMenu("Fichier")
        menu_fichier.addAction("📤 Exporter vers Excel AGO", self.exporter_excel)
        menu_aide = menubar.addMenu("❓ Aide")
        menu_aide.addAction("Guide d'utilisation", self.ouvrir_aide)
        menu_aide.addSeparator()
        menu_aide.addAction("À propos", self.ouvrir_apropos)

        # Barre d'outils
        toolbar = QHBoxLayout()

        self.btn_import = QPushButton("📥 1. Importer Export Portail (.xlsx)")
        self.btn_import.clicked.connect(self.importer)
        self.btn_import.setStyleSheet("font-weight: bold; padding: 8px; background-color: #4CAF50; color: white;")
        self.btn_import.setToolTip(
            "Importer un export Excel depuis le portail AGO.\n"
            "Le fichier doit contenir l'onglet « Liste des evenements ».\n"
            "Seuls les DPS incomplets (Présents < Requis) sont chargés.")
        toolbar.addWidget(self.btn_import)

        self.lbl_semaine = QLabel("Semaine active :")
        self.lbl_semaine.setStyleSheet("font-weight: bold; font-size: 14px; margin-left: 20px; color: #2196F3;")
        toolbar.addWidget(self.lbl_semaine)

        self.combo_semaine = QComboBox()
        self.combo_semaine.setStyleSheet("font-weight: bold; font-size: 14px; padding: 4px;")
        self.combo_semaine.setToolTip("Sélectionner la semaine à afficher.\nLes semaines sont conservées entre les sessions.")
        self.combo_semaine.currentIndexChanged.connect(self.on_semaine_changed)
        toolbar.addWidget(self.combo_semaine)

        self.btn_nouvelle_semaine = QPushButton("📅 Nouvelle semaine")
        self.btn_nouvelle_semaine.clicked.connect(self.creer_semaine)
        self.btn_nouvelle_semaine.setStyleSheet("font-weight: bold; padding: 8px; background-color: #2196F3; color: white;")
        self.btn_nouvelle_semaine.setToolTip("Créer une nouvelle semaine vide sans importer de fichier.")
        toolbar.addWidget(self.btn_nouvelle_semaine)

        toolbar.addStretch()

        self.btn_reinjecter = QPushButton("🔄 Réinjecter les renforts")
        self.btn_reinjecter.clicked.connect(self.reinjecter_renforts)
        self.btn_reinjecter.setEnabled(False)
        self.btn_reinjecter.setStyleSheet("font-weight: bold; padding: 8px; background-color: #E65100; color: white;")
        self.btn_reinjecter.setToolTip(
            "Restaurer les renforts ajoutés manuellement avant le dernier ré-import.\n"
            "Ce bouton s'active automatiquement si des renforts manuels ont été détectés\n"
            "lors du dernier import de la semaine active.")
        toolbar.addWidget(self.btn_reinjecter)

        self.btn_synthese = QPushButton("📊 Synthèse de la semaine")
        self.btn_synthese.clicked.connect(self.synthese_semaine)
        self.btn_synthese.setEnabled(False)
        self.btn_synthese.setStyleSheet("font-weight: bold; padding: 8px; background-color: #7B1FA2; color: white;")
        self.btn_synthese.setToolTip("Afficher un récapitulatif complet de la semaine active.\nPermet de copier pour WhatsApp ou d'imprimer.")
        toolbar.addWidget(self.btn_synthese)

        layout.addLayout(toolbar)

        # Zone tableau : colonne Antenne figée à gauche + tableau jours scrollable
        table_container = QWidget()
        table_h = QHBoxLayout(table_container)
        table_h.setSpacing(0)
        table_h.setContentsMargins(0, 0, 0, 0)

        # --- Colonne Antenne figée ---
        self.frozen_table = QTableWidget()
        self.frozen_table.setColumnCount(1)
        self.frozen_table.setHorizontalHeaderLabels(["Antenne"])
        self.frozen_table.horizontalHeader().setFixedHeight(60)
        self.frozen_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.frozen_table.setColumnWidth(0, 150)
        self.frozen_table.setFixedWidth(152)
        self.frozen_table.verticalHeader().setVisible(False)
        self.frozen_table.verticalHeader().setDefaultSectionSize(35)
        self.frozen_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.frozen_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.frozen_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.frozen_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.frozen_table.setStyleSheet("""
            QTableWidget { border: none; border-right: 2px solid #1E3C72; }
            QHeaderView::section {
                background-color: #1E3C72;
                color: white;
                font-weight: bold;
                font-size: 11px;
                border: none;
            }
        """)
        table_h.addWidget(self.frozen_table)

        # --- Tableau jours ---
        self.table = QTableWidget()
        nb_cols = len(JOURS_SEMAINE) * 4
        self.table.setColumnCount(nb_cols)

        header = TwoLineHeaderView(self.table)
        self.table.setHorizontalHeader(header)

        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(35)

        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)

        for i in range(nb_cols):
            self.table.setColumnWidth(i, 120)

        self.table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.table.cellChanged.connect(self.on_cell_changed)

        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        table_h.addWidget(self.table)

        # Synchronisation du défilement vertical
        self.table.verticalScrollBar().valueChanged.connect(
            self.frozen_table.verticalScrollBar().setValue)
        self.frozen_table.verticalScrollBar().valueChanged.connect(
            self.table.verticalScrollBar().setValue)

        layout.addWidget(table_container)
        self.statusBar().showMessage("Prêt. Cliquez sur 'Importer Export Portail' pour commencer.")
        self.refresh_semaine_selector()

    # ------------------------------------------------------------------ import

    def importer(self):
        downloads = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DownloadLocation)
        path, _ = QFileDialog.getOpenFileName(self, "Sélectionner l'export portail", downloads, "Fichiers Excel (*.xlsx)")
        if not path:
            return
        try:
            imported_week, df, resume = importer_evenements(path)
            if imported_week is None:
                QMessageBox.warning(self, "Import — Aucun DPS incomplet", resume)
                return
            self.refresh_semaine_selector(select_week=imported_week)
            self.statusBar().showMessage(f"Import réussi ! Semaine {imported_week} chargée.")
            QMessageBox.information(self, f"Import réussi — Semaine {imported_week}", resume)
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de lire l'export :\n{e}")

    # --------------------------------------------------------- sélecteur semaine

    def creer_semaine(self):
        import datetime as dt
        from PyQt6.QtWidgets import QDialog, QFormLayout, QSpinBox, QLineEdit, QDialogButtonBox

        today  = dt.date.today()
        monday = today - dt.timedelta(days=today.weekday())
        sunday = monday + dt.timedelta(days=6)
        week_n = today.isocalendar()[1]

        dlg = QDialog(self)
        dlg.setWindowTitle("Nouvelle semaine")
        dlg.setFixedWidth(340)
        form = QFormLayout(dlg)

        spin = QSpinBox()
        spin.setRange(1, 53)
        spin.setValue(week_n)

        edit_deb = QLineEdit(monday.strftime('%Y-%m-%d'))
        edit_fin = QLineEdit(sunday.strftime('%Y-%m-%d'))

        form.addRow("Numéro de semaine :", spin)
        form.addRow("Date début (YYYY-MM-DD) :", edit_deb)
        form.addRow("Date fin (YYYY-MM-DD) :", edit_fin)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        form.addRow(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        numero = spin.value()
        date_debut = edit_deb.text().strip()
        date_fin   = edit_fin.text().strip()

        try:
            conn = get_conn()
            c = conn.cursor()
            c.execute(
                "INSERT INTO semaines (numero, date_debut, date_fin) VALUES (%s, %s, %s)"
                " ON CONFLICT (numero) DO NOTHING",
                (numero, date_debut, date_fin)
            )
            conn.commit()
            conn.close()
            self.refresh_semaine_selector(select_week=numero)
            self.statusBar().showMessage(f"Semaine {numero} créée.")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de créer la semaine :\n{e}")

    def refresh_semaine_selector(self, select_week=None):
        self.combo_semaine.blockSignals(True)
        self.combo_semaine.clear()

        conn = get_conn()
        c = conn.cursor()
        c.execute("SELECT numero FROM semaines ORDER BY numero DESC")
        weeks = [r[0] for r in c.fetchall()]
        conn.close()

        for w in weeks:
            self.combo_semaine.addItem(f"Semaine {w}", w)

        if select_week is not None:
            idx = next((i for i in range(self.combo_semaine.count())
                        if self.combo_semaine.itemData(i) == select_week), -1)
            if idx == -1:
                self.combo_semaine.addItem(f"Semaine {select_week}", select_week)
                idx = self.combo_semaine.count() - 1
            self.combo_semaine.setCurrentIndex(idx)
            self.current_semaine = select_week
        elif weeks:
            self.combo_semaine.setCurrentIndex(0)
            self.current_semaine = weeks[0]
        else:
            self.current_semaine = None

        self.combo_semaine.blockSignals(False)

        if self.current_semaine is not None:
            self.btn_synthese.setEnabled(True)
            self.charger_vue()
        else:
            self.btn_synthese.setEnabled(False)
            self.table.setRowCount(0)
            self.frozen_table.setRowCount(0)
        self._refresh_btn_reinjecter()

    def on_semaine_changed(self, index):
        if index < 0:
            return
        self.current_semaine = self.combo_semaine.itemData(index)
        self.charger_vue()
        self._refresh_btn_reinjecter()

    # ---------------------------------------------------------------- vue tableau

    def charger_vue(self):
        if not self.current_semaine:
            return

        conn = get_conn()
        c = conn.cursor()
        c.execute("SELECT id FROM semaines WHERE numero = %s", (self.current_semaine,))
        res = c.fetchone()
        if not res:
            conn.close()
            return
        semaine_id = res[0]

        c.execute("""SELECT antenne, jour, nom_dps, nb, tl, id, est_renfort, parent_dps_id
                     FROM dps WHERE semaine_id = %s ORDER BY antenne, jour, est_renfort""", (semaine_id,))
        rows = c.fetchall()
        conn.close()

        # Somme des IS de renforts engagés par DPS parent
        renfort_par_parent = {}
        for r in rows:
            if r[6] and r[7]:  # est_renfort et parent_dps_id non nuls
                renfort_par_parent[r[7]] = renfort_par_parent.get(r[7], 0) + r[3]

        # Lookup id DPS principal → (effective_nb, tl) pour corriger le Manque des renforts
        principal_par_id = {}
        for r in rows:
            if not r[6]:  # est_renfort = 0
                eff = r[3] + renfort_par_parent.get(r[5], 0)
                principal_par_id[r[5]] = (eff, r[4])

        # Grouper par (antenne, jour)
        data = {}
        for r in rows:
            key = (r[0], r[1])
            if key not in data:
                data[key] = {"principaux": [], "renforts": []}
            entry = {"nom": r[2], "nb": r[3], "tl": r[4], "id": r[5], "est_renfort": r[6], "parent_id": r[7]}
            if r[6]:
                data[key]["renforts"].append(entry)
            else:
                data[key]["principaux"].append(entry)

        # Nombre de lignes par antenne
        lignes_par_antenne = {}
        for antenne in ANTENNES_ORDRE:
            max_lignes = 1
            for jour in JOURS_SEMAINE:
                key = (antenne, jour)
                if key in data:
                    nb = len(data[key]["principaux"]) + len(data[key]["renforts"])
                    if nb > max_lignes:
                        max_lignes = nb
            lignes_par_antenne[antenne] = max_lignes

        total_lignes = sum(lignes_par_antenne.values())

        self.table.blockSignals(True)
        self.table.setRowCount(0)
        self.frozen_table.setRowCount(0)
        self.table.setRowCount(total_lignes)
        self.frozen_table.setRowCount(total_lignes)

        row_index = 0
        for antenne in ANTENNES_ORDRE:
            nb_lignes = lignes_par_antenne[antenne]

            for ligne_idx in range(nb_lignes):
                # Colonne Antenne figée
                ant_item = QTableWidgetItem(antenne)
                ant_item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                self.frozen_table.setItem(row_index, 0, ant_item)
                self.frozen_table.setRowHeight(row_index, 35)

                for j_idx, jour in enumerate(JOURS_SEMAINE):
                    col_offset = j_idx * 4
                    key = (antenne, jour)

                    if key not in data:
                        row_index_tmp = row_index  # just continue
                        continue

                    # DPS principaux
                    if ligne_idx < len(data[key]["principaux"]):
                        d = data[key]["principaux"][ligne_idx]
                        renforts_nb = renfort_par_parent.get(d["id"], 0)
                        effective_nb = d["nb"] + renforts_nb
                        manque = max(0, d["tl"] - effective_nb)

                        nom_item = QTableWidgetItem(d["nom"])
                        nom_item.setData(_DPS_ID_ROLE, d["id"])
                        self.table.setItem(row_index, col_offset, nom_item)

                        # Engagés : valeur calculée, non éditable pour les parents
                        eng_item = QTableWidgetItem(str(effective_nb))
                        eng_item.setFlags(eng_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        self.table.setItem(row_index, col_offset + 1, eng_item)

                        self.table.setItem(row_index, col_offset + 2, QTableWidgetItem(str(manque)))
                        self.table.setItem(row_index, col_offset + 3, QTableWidgetItem(str(d["tl"])))
                        self._colorer_cellule(row_index, col_offset, effective_nb, d["tl"])

                    # Lignes de renforts [R]
                    elif ligne_idx < len(data[key]["principaux"]) + len(data[key]["renforts"]):
                        r_idx = ligne_idx - len(data[key]["principaux"])
                        r = data[key]["renforts"][r_idx]

                        nom_item = QTableWidgetItem(r["nom"])
                        nom_item.setData(_DPS_ID_ROLE, r["id"])
                        self.table.setItem(row_index, col_offset, nom_item)

                        self.table.setItem(row_index, col_offset + 1, QTableWidgetItem(str(r["nb"])))

                        parent_info = principal_par_id.get(r.get("parent_id"))
                        if parent_info:
                            p_eff, p_tl = parent_info
                            manque = max(0, p_tl - p_eff)
                            self._colorer_cellule(row_index, col_offset, p_eff, p_tl, est_renfort=True)
                        else:
                            manque = max(0, r["tl"] - r["nb"])
                            self._colorer_cellule(row_index, col_offset, r["nb"], r["tl"], est_renfort=True)

                        self.table.setItem(row_index, col_offset + 2, QTableWidgetItem(str(manque)))
                        self.table.setItem(row_index, col_offset + 3, QTableWidgetItem(str(r["tl"])))

                row_index += 1

        self.table.blockSignals(False)

    def _colorer_cellule(self, row, col_start, nb, tl, est_renfort=False):
        if est_renfort:
            color = QColor(COLOR_RENFORT_KO) if nb < tl else QColor(COLOR_RENFORT_OK)
        else:
            color = QColor(COLOR_ORANGE) if nb < tl else QColor(COLOR_GREEN)
        for i in range(col_start, col_start + 4):
            item = self.table.item(row, i)
            if item:
                item.setBackground(color)

    # --------------------------------------------------------- édition en live

    def on_cell_changed(self, row, col):
        col_in_day = col % 4
        if col_in_day not in (1, 3):  # Engagés ou Σ uniquement
            return

        jour_idx = col // 4
        col_offset = jour_idx * 4

        dps_item = self.table.item(row, col_offset)
        if not dps_item:
            return
        dps_id = dps_item.data(_DPS_ID_ROLE)
        if not dps_id:
            return

        changed_item = self.table.item(row, col)
        if not changed_item:
            return

        text = changed_item.text().strip()
        if not text.isdigit():
            return
        val = int(text)

        conn = get_conn()
        c = conn.cursor()
        if col_in_day == 1:
            c.execute("UPDATE dps SET nb = %s WHERE id = %s", (val, dps_id))
        else:
            c.execute("UPDATE dps SET tl = %s WHERE id = %s", (val, dps_id))
        conn.commit()
        conn.close()

        scroll_v = self.table.verticalScrollBar().value()
        scroll_h = self.table.horizontalScrollBar().value()
        self.charger_vue()
        self.table.verticalScrollBar().setValue(scroll_v)
        self.table.horizontalScrollBar().setValue(scroll_h)

    # ----------------------------------------------------------- menu contextuel

    def show_context_menu(self, pos):
        row = self.table.rowAt(pos.y())
        col = self.table.columnAt(pos.x())
        if row < 0:
            return

        antenne_item = self.frozen_table.item(row, 0)
        if not antenne_item or antenne_item.text() not in ANTENNES_ORDRE:
            return

        jour_idx = col // 4
        col_offset = jour_idx * 4
        jour = JOURS_SEMAINE[jour_idx]

        item = self.table.item(row, col_offset)
        if not item or not item.text():
            return

        nom_dps = item.text()
        dps_id = item.data(_DPS_ID_ROLE)

        nb_item = self.table.item(row, col_offset + 1)
        tl_item = self.table.item(row, col_offset + 3)
        nb = int(nb_item.text()) if nb_item and nb_item.text().isdigit() else 0
        tl = int(tl_item.text()) if tl_item and tl_item.text().isdigit() else 0

        menu = QMenu(self)
        if nb < tl and not nom_dps.startswith("[R]"):
            action_renfort = menu.addAction(f"💡 Proposer des renforts pour : {nom_dps}")
            action_renfort.triggered.connect(
                lambda: self.ouvrir_renfort(antenne_item.text(), jour, nom_dps, nb, tl, dps_id))

        action_supprimer = menu.addAction("🗑️ Supprimer ce DPS")
        action_supprimer.triggered.connect(lambda: self.supprimer_dps(dps_id))
        menu.exec(self.table.mapToGlobal(pos))

    # ----------------------------------------------------------------- renforts

    def ouvrir_renfort(self, antenne, jour, nom_dps, nb, tl, dps_id):
        besoin = tl - nb

        conn = get_conn()
        c = conn.cursor()
        c.execute("SELECT id FROM semaines WHERE numero = %s", (self.current_semaine,))
        res = c.fetchone()
        conn.close()

        if not res:
            QMessageBox.critical(self, "Erreur", "Semaine introuvable en base.")
            return

        semaine_id = res[0]
        suggestions = suggerer_renforts(antenne, jour, semaine_id, besoin)
        all_suggestions = toutes_disponibilites(antenne, jour, semaine_id)
        dps_info = {'antenne': antenne, 'jour': jour, 'nom_dps': nom_dps, 'nb': nb, 'tl': tl}
        popup = RenfortPopup(dps_info, suggestions, all_suggestions, self)

        if popup.exec() == 1:
            selections = popup.get_selected_antennes()
            if not selections:
                return

            besoin_total = tl - nb
            nb_antennes = len(selections)
            base = besoin_total // nb_antennes
            reste = besoin_total % nb_antennes

            try:
                for i, (antenne_cible, capacite) in enumerate(selections):
                    nb_a_envoyer = base + (1 if i < reste else 0)
                    ajouter_renforts(semaine_id, dps_id, antenne_cible, jour, nom_dps, nb_a_envoyer)
                QMessageBox.information(self, "Succès", "Les lignes [R] ont été ajoutées en base.")
                self.charger_vue()
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Échec de l'ajout des renforts :\n{e}")

    def supprimer_dps(self, dps_id):
        if not dps_id:
            QMessageBox.warning(self, "Attention", "Impossible de supprimer ce DPS.")
            return
        reply = QMessageBox.question(self, "Confirmation", "Voulez-vous vraiment supprimer ce DPS ?")
        if reply == QMessageBox.StandardButton.Yes:
            conn = get_conn()
            c = conn.cursor()
            c.execute("DELETE FROM dps WHERE parent_dps_id = %s", (dps_id,))
            c.execute("DELETE FROM dps WHERE id = %s", (dps_id,))
            conn.commit()
            conn.close()
            self.charger_vue()
            self.statusBar().showMessage("DPS supprimé.")

    # -------------------------------------------------- réinjection renforts manuels

    def _refresh_btn_reinjecter(self):
        if not self.current_semaine:
            self.btn_reinjecter.setEnabled(False)
            self.btn_reinjecter.setText("🔄 Réinjecter les renforts")
            return
        conn = get_conn()
        c = conn.cursor()
        try:
            c.execute("SELECT COUNT(*) FROM renforts_backup WHERE semaine_num = %s", (self.current_semaine,))
            count = c.fetchone()[0]
        except Exception:
            count = 0
        conn.close()
        if count > 0:
            s = "s" if count > 1 else ""
            self.btn_reinjecter.setEnabled(True)
            self.btn_reinjecter.setText(f"🔄 Réinjecter ({count} renfort{s})")
        else:
            self.btn_reinjecter.setEnabled(False)
            self.btn_reinjecter.setText("🔄 Réinjecter les renforts")

    def reinjecter_renforts(self):
        if not self.current_semaine:
            return
        conn = get_conn()
        c = conn.cursor()
        c.execute("SELECT id FROM semaines WHERE numero = %s", (self.current_semaine,))
        res = c.fetchone()
        if not res:
            conn.close()
            return
        semaine_id = res[0]

        c.execute("""SELECT antenne, jour, nom_dps, nb, tl, parent_antenne, parent_jour, parent_nom
                     FROM renforts_backup WHERE semaine_num = %s""", (self.current_semaine,))
        backup = c.fetchall()

        if not backup:
            QMessageBox.information(self, "Réinjection", "Aucun renfort manuel sauvegardé pour cette semaine.")
            conn.close()
            return

        injectes = 0
        deja_presents = []
        introuvables = []

        for antenne, jour, nom_dps, nb, tl, parent_ant, parent_jour, parent_nom in backup:
            # Vérifier si un renfort identique existe déjà (doublon issu du nouvel import)
            c.execute("""SELECT id FROM dps
                         WHERE semaine_id=%s AND antenne=%s AND jour=%s AND nom_dps=%s AND est_renfort=1""",
                      (semaine_id, antenne, jour, nom_dps))
            if c.fetchone():
                deja_presents.append(f"{nom_dps}  ({antenne} / {jour})")
                continue

            parent_db_id = None
            if parent_ant and parent_jour and parent_nom:
                c.execute("""SELECT id FROM dps
                             WHERE semaine_id=%s AND antenne=%s AND jour=%s AND nom_dps=%s AND est_renfort=0""",
                          (semaine_id, parent_ant, parent_jour, parent_nom))
                res_p = c.fetchone()
                if res_p:
                    parent_db_id = res_p[0]

            if parent_db_id is not None:
                c.execute("""INSERT INTO dps (semaine_id, antenne, jour, nom_dps, nb, tl, est_renfort, parent_dps_id, est_manuel)
                             VALUES (%s, %s, %s, %s, %s, %s, 1, %s, 1)""",
                          (semaine_id, antenne, jour, nom_dps, nb, tl, parent_db_id))
                injectes += 1
            else:
                introuvables.append(f"{nom_dps}  →  parent : {parent_ant} / {parent_jour}")

        conn.commit()
        c.execute("DELETE FROM renforts_backup WHERE semaine_num = %s", (self.current_semaine,))
        conn.commit()
        conn.close()

        lignes = [f"✅ {injectes} renfort(s) réinjecté(s) avec succès."]
        if deja_presents:
            lignes.append(f"\nℹ️ {len(deja_presents)} renfort(s) déjà présent(s) dans l'import (ignorés) :")
            lignes.extend(f"  • {l}" for l in deja_presents)
        if introuvables:
            lignes.append(f"\n⚠️ {len(introuvables)} renfort(s) non retrouvé(s)\n(DPS parent introuvable après ré-import) :")
            lignes.extend(f"  • {l}" for l in introuvables)

        QMessageBox.information(self, "Réinjection des renforts manuels", "\n".join(lignes))
        self.charger_vue()
        self._refresh_btn_reinjecter()

    # ------------------------------------------------------------------- export

    def exporter_excel(self):
        if not self.current_semaine:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter vers Excel", f"AGO-Suivi-S{self.current_semaine}.xlsx", "Fichiers Excel (*.xlsx)")
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = f"Semaine {self.current_semaine}"

            # En-têtes 2 lignes
            ws.merge_cells('A1:A2')
            ws['A1'] = 'Antenne'
            ws['A1'].font = Font(bold=True, color='FFFFFF', size=11)
            ws['A1'].fill = PatternFill(start_color='1E3C72', end_color='1E3C72', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            for i, jour in enumerate(JOURS_SEMAINE):
                col_debut = 2 + i * 4
                ws.merge_cells(start_row=1, start_column=col_debut, end_row=1, end_column=col_debut + 3)
                cell = ws.cell(row=1, column=col_debut, value=jour)
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.fill = PatternFill(start_color='1E3C72', end_color='1E3C72', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

                for j, sous_col in enumerate(['DPS', 'Engagés', 'Manque', 'Σ']):
                    c = ws.cell(row=2, column=col_debut + j, value=sous_col)
                    c.font = Font(bold=True, color='FFFFFF', size=9)
                    c.fill = PatternFill(start_color='1E3C72', end_color='1E3C72', fill_type='solid')
                    c.alignment = Alignment(horizontal='center', vertical='center')

            thin = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
            for r in ws[1:2]:
                for cell in r:
                    cell.border = thin

            # Données depuis la DB
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT id FROM semaines WHERE numero = %s", (self.current_semaine,))
            semaine_id = cur.fetchone()[0]
            cur.execute("""SELECT antenne, jour, nom_dps, nb, tl, id, est_renfort, parent_dps_id
                           FROM dps WHERE semaine_id = %s ORDER BY antenne, jour, est_renfort""", (semaine_id,))
            db_rows = cur.fetchall()
            conn.close()

            renfort_par_parent = {}
            for r in db_rows:
                if r[6] and r[7]:
                    renfort_par_parent[r[7]] = renfort_par_parent.get(r[7], 0) + r[3]

            principal_par_id = {}
            for r in db_rows:
                if not r[6]:
                    eff = r[3] + renfort_par_parent.get(r[5], 0)
                    principal_par_id[r[5]] = (eff, r[4])

            data = {}
            for r in db_rows:
                key = (r[0], r[1])
                if key not in data:
                    data[key] = []
                data[key].append({"nom": r[2], "nb": r[3], "tl": r[4], "id": r[5], "est_renfort": r[6], "parent_id": r[7]})

            row_idx = 3
            for antenne in ANTENNES_ORDRE:
                max_lignes = max(
                    (len(data.get((antenne, j), [])) for j in JOURS_SEMAINE), default=1)
                max_lignes = max(max_lignes, 1)

                for ligne_idx in range(max_lignes):
                    ws.cell(row=row_idx, column=1, value=antenne).border = thin

                    for j_idx, jour in enumerate(JOURS_SEMAINE):
                        col_offset = 2 + j_idx * 4
                        key = (antenne, jour)
                        if key in data and ligne_idx < len(data[key]):
                            d = data[key][ligne_idx]
                            if d["est_renfort"]:
                                effective_nb = d["nb"]
                                parent_info = principal_par_id.get(d.get("parent_id"))
                                if parent_info:
                                    p_eff, p_tl = parent_info
                                    manque = max(0, p_tl - p_eff)
                                    color = '9DC3E6' if p_eff >= p_tl else 'FFE699'
                                else:
                                    manque = max(0, d["tl"] - effective_nb)
                                    color = '9DC3E6' if effective_nb >= d["tl"] else 'FFE699'
                            else:
                                effective_nb = d["nb"] + renfort_par_parent.get(d["id"], 0)
                                manque = max(0, d["tl"] - effective_nb)
                                color = 'FFC000' if effective_nb < d["tl"] else 'C6EFCE'

                            ws.cell(row=row_idx, column=col_offset, value=d["nom"])
                            ws.cell(row=row_idx, column=col_offset + 1, value=effective_nb)
                            ws.cell(row=row_idx, column=col_offset + 2, value=manque)
                            ws.cell(row=row_idx, column=col_offset + 3, value=d["tl"])
                            for col in range(col_offset, col_offset + 4):
                                ws.cell(row=row_idx, column=col).fill = PatternFill(
                                    start_color=color, end_color=color, fill_type='solid')

                        for col in range(col_offset, col_offset + 4):
                            ws.cell(row=row_idx, column=col).border = thin

                    row_idx += 1

            wb.save(path)
            QMessageBox.information(self, "Succès", f"Fichier exporté :\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible d'exporter :\n{e}")

    # -------------------------------------------------------------------- aide

    def ouvrir_aide(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Guide d'utilisation — PHARE")
        dlg.resize(620, 520)
        layout = QVBoxLayout(dlg)

        browser = QTextBrowser()
        browser.setOpenExternalLinks(False)
        browser.setHtml("""
<style>
  body  { font-family: Arial, sans-serif; font-size: 13px; margin: 12px; }
  h2    { color: #1E3C72; border-bottom: 2px solid #1E3C72; padding-bottom: 4px; }
  h3    { color: #2196F3; margin-top: 14px; }
  ul    { margin: 4px 0 8px 18px; }
  li    { margin-bottom: 3px; }
  .tag  { display: inline-block; border-radius: 3px; padding: 1px 6px;
          font-size: 11px; font-weight: bold; }
  .or   { background:#FFC000; }
  .ve   { background:#C6EFCE; }
  .bl   { background:#9DC3E6; }
  .ja   { background:#FFE699; }
</style>

<h2>PHARE — Suivi Départemental des DPS</h2>
<p style="color:#555;font-size:11px"><i>Plateforme Harmonisée d'Affectation des Renforts et des Effectifs</i></p>

<h3>0. Données partagées entre postes</h3>
<ul>
  <li>Les données sont <b>hébergées dans le cloud</b> (Supabase) et communes
      à tous les postes de l'antenne.</li>
  <li>Chaque sauvegarde (import, saisie, renfort) est immédiatement visible
      depuis n'importe quel poste connecté.</li>
  <li>Une <b>connexion internet</b> est nécessaire. En cas de perte de connexion,
      l'application affichera une erreur au prochain accès.</li>
  <li>Pour voir les modifications faites par un collègue, changez de semaine
      et revenez, ou relancez l'application.</li>
</ul>

<h3>1. Importer un fichier</h3>
<ul>
  <li>Cliquez sur <b>📥 1. Importer Export Portail (.xlsx)</b>.</li>
  <li>Le fichier doit contenir l'onglet <b>« Liste des evenements »</b>.</li>
  <li>La colonne A (numéro d'activité) est utilisée pour lier chaque DPS
      à ses renforts existants.</li>
  <li>Seuls les groupes <b>incomplets</b> (Présents &lt; Requis) sont chargés.</li>
  <li>Un nouvel import de la même semaine <b>remplace</b> les données existantes
      pour cette semaine (les autres semaines ne sont pas touchées).</li>
</ul>

<h3>2. Navigation</h3>
<ul>
  <li>Utilisez le sélecteur <b>Semaine active</b> pour changer de semaine.</li>
  <li>La colonne <b>Antenne</b> à gauche reste figée lors du défilement horizontal.</li>
  <li>Les données sont <b>conservées</b> entre les sessions (base SQLite locale).</li>
</ul>

<h3>3. Lire le tableau</h3>
<ul>
  <li><b>DPS</b> — Nom de l'activité (avec horaires)</li>
  <li><b>Engagés</b> — IS confirmés (renforts inclus pour les DPS principaux)</li>
  <li><b>Manque</b> — IS encore nécessaires</li>
  <li><b>Σ</b> — Effectif total requis</li>
</ul>
<p>
  <span class="tag or">Orange</span> DPS principal incomplet &nbsp;
  <span class="tag ve">Vert</span> DPS principal complet &nbsp;
  <span class="tag ja">Jaune</span> Renfort [R] incomplet &nbsp;
  <span class="tag bl">Bleu</span> Renfort [R] complet
</p>

<h3>4. Modifier manuellement</h3>
<ul>
  <li><b>Double-clic</b> sur la colonne <i>Engagés</i> d'un <b>[R]</b> pour
      saisir le nombre d'IS engagés.</li>
  <li><b>Double-clic</b> sur <i>Σ</i> pour modifier le besoin total d'une ligne.</li>
  <li>La sauvegarde est <b>immédiate</b> en base ; le Manque du DPS parent
      se recalcule automatiquement.</li>
</ul>

<h3>5. Gérer les renforts</h3>
<ul>
  <li><b>Clic droit</b> sur un DPS incomplet → <i>Proposer des renforts</i>.</li>
  <li>Les 3 antennes voisines disponibles sont affichées par défaut.</li>
  <li>Cochez <b>🌐 Afficher toutes les antennes</b> pour voir toutes les options.</li>
  <li>Sélectionnez une ou plusieurs antennes et cliquez <b>Valider</b> :
      une ligne <b>[R]</b> est créée avec Engagés = 0 (à remplir manuellement).</li>
</ul>

<h3>6. Supprimer un DPS</h3>
<ul>
  <li><b>Clic droit</b> → <i>Supprimer ce DPS</i> supprime la ligne
      et tous ses renforts [R] associés.</li>
</ul>

<h3>7. Réinjecter les renforts manuels</h3>
<ul>
  <li>Lors d'un <b>ré-import</b> d'une semaine, les renforts ajoutés manuellement
      (via <i>Proposer des renforts</i>) sont sauvegardés automatiquement.</li>
  <li>Le bouton <b>🔄 Réinjecter</b> s'active dans la barre d'outils et indique
      le nombre de renforts en attente.</li>
  <li>Cliquez dessus pour les restaurer. L'appli vérifie les doublons :
      un renfort déjà présent dans le nouvel import ne sera pas créé en double.</li>
  <li>Si le DPS parent a changé de nom côté portail, le renfort est signalé
      comme <i>non retrouvé</i> dans le résumé.</li>
</ul>

<h3>8. Synthèse de la semaine</h3>
<ul>
  <li>Cliquez sur <b>📊 Synthèse de la semaine</b> pour un récapitulatif complet.</li>
  <li><b>📱 Copier pour WhatsApp</b> — génère un texte formaté (gras, italique)
      prêt à coller dans une conversation.</li>
  <li><b>🖨️ Imprimer</b> — envoie la synthèse vers l'imprimante ou PDF.</li>
</ul>

<h3>9. Exporter vers Excel</h3>
<ul>
  <li>Menu <b>Fichier → 📤 Exporter vers Excel AGO</b> pour générer
      un fichier .xlsx formaté avec couleurs et en-têtes sur 2 niveaux.</li>
</ul>
""")
        layout.addWidget(browser)

        btn = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        btn.rejected.connect(dlg.reject)
        layout.addWidget(btn)
        dlg.exec()

    def synthese_semaine(self):
        if not self.current_semaine:
            return

        conn = get_conn()
        c = conn.cursor()
        c.execute("SELECT id, date_debut, date_fin FROM semaines WHERE numero = %s", (self.current_semaine,))
        sem = c.fetchone()
        if not sem:
            conn.close()
            QMessageBox.information(self, "Synthèse", "Aucune donnée pour cette semaine.")
            return
        semaine_id, date_debut, date_fin = sem

        c.execute("""SELECT antenne, jour, nom_dps, nb, tl, id, est_renfort, parent_dps_id
                     FROM dps WHERE semaine_id = %s ORDER BY antenne, jour, est_renfort""", (semaine_id,))
        rows = c.fetchall()
        conn.close()

        if not rows:
            QMessageBox.information(self, "Synthèse", "Aucun DPS enregistré pour la semaine %s." % self.current_semaine)
            return

        # Structures de données
        renfort_par_parent = {}
        for r in rows:
            if r[6] and r[7]:
                renfort_par_parent[r[7]] = renfort_par_parent.get(r[7], 0) + r[3]

        renforts_de = {}  # parent_id -> [{antenne, nb, tl}]
        data = {}
        for r in rows:
            ant, jour, nom, nb, tl, did, est_r, parent_id = r
            data.setdefault(ant, {}).setdefault(jour, {'principaux': [], 'renforts': []})
            entry = {'nom': nom, 'nb': nb, 'tl': tl, 'id': did, 'antenne': ant}
            if est_r:
                data[ant][jour]['renforts'].append(entry)
                if parent_id:
                    renforts_de.setdefault(parent_id, []).append({'antenne': ant, 'nb': nb, 'tl': tl})
            else:
                data[ant][jour]['principaux'].append(entry)

        # Statistiques globales
        principaux = [r for r in rows if not r[6]]
        renforts   = [r for r in rows if r[6]]
        total_tl   = sum(r[4] for r in principaux)
        total_eng  = sum(r[3] + renfort_par_parent.get(r[5], 0) for r in principaux)
        total_man  = sum(max(0, r[4] - (r[3] + renfort_par_parent.get(r[5], 0))) for r in principaux)
        nb_incomplets = sum(1 for r in principaux if r[3] + renfort_par_parent.get(r[5], 0) < r[4])

        def fmt_date(d):
            try:
                from datetime import datetime
                return datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m/%Y')
            except Exception:
                return d

        periode = f"Du {fmt_date(date_debut)} au {fmt_date(date_fin)}"

        # ── Génération HTML ──────────────────────────────────────────
        html_parts = [f"""<style>
body{{font-family:Arial,sans-serif;font-size:13px;margin:10px}}
h2{{color:#1E3C72;border-bottom:2px solid #1E3C72;padding-bottom:4px}}
h3{{color:#2196F3;margin:14px 0 4px}}
h4{{margin:8px 0 2px;color:#333}}
table{{border-collapse:collapse;width:100%;margin:4px 0}}
td,th{{border:1px solid #ddd;padding:4px 8px;font-size:12px}}
th{{background:#1E3C72;color:#fff}}
.ok{{background:#C6EFCE}}.ko{{background:#FFC000}}
.rok{{background:#9DC3E6}}.rko{{background:#FFE699}}
.bilan{{background:#f0f4ff;border:1px solid #1E3C72;padding:10px;border-radius:6px;margin-top:12px}}
</style>
<h2>📊 Synthèse — Semaine {self.current_semaine}</h2>
<p><i>{periode}</i></p>"""]

        wa_parts = [f"📊 *SYNTHÈSE — SEMAINE {self.current_semaine}*",
                    f"_{periode}_", ""]

        for antenne in ANTENNES_ORDRE:
            if antenne not in data:
                continue
            jours_ant = {j: v for j, v in data[antenne].items()
                         if v['principaux'] or v['renforts']}
            if not jours_ant:
                continue

            html_parts.append(f"<h3>{antenne}</h3>")
            wa_parts.append(f"*{antenne}*")

            for jour in JOURS_SEMAINE:
                if jour not in jours_ant:
                    continue
                grp = jours_ant[jour]
                html_parts.append(f"<h4>📅 {jour}</h4><table><tr><th>DPS</th><th>Engagés</th><th>Σ</th><th>Manque</th><th>Statut</th></tr>")
                wa_parts.append(f"  📅 {jour}")

                for d in grp['principaux']:
                    eff = d['nb'] + renfort_par_parent.get(d['id'], 0)
                    man = max(0, d['tl'] - eff)
                    css = 'ok' if eff >= d['tl'] else 'ko'
                    ic  = '✅' if eff >= d['tl'] else '⚠️'
                    html_parts.append(
                        f"<tr class='{css}'><td>{d['nom']}</td><td>{eff}</td>"
                        f"<td>{d['tl']}</td><td>{man}</td><td>{ic}</td></tr>")
                    wa_parts.append(f"  {ic} {d['nom']}")
                    wa_parts.append(f"     Engagés : {eff}/{d['tl']}" + (f" · Manque : {man}" if man else " · Complet"))

                    for rf in renforts_de.get(d['id'], []):
                        parent_ok = eff >= d['tl']
                        ric = '✅' if parent_ok else '⏳'
                        html_parts.append(
                            f"<tr class='{'rok' if parent_ok else 'rko'}'>"
                            f"<td>&nbsp;&nbsp;↳ [R] {rf['antenne']}</td>"
                            f"<td>{rf['nb']}</td><td>{d['tl']}</td>"
                            f"<td>{man}</td><td>{ric}</td></tr>")
                        wa_parts.append(f"     ↳ [R] {rf['antenne']} : {rf['nb']}/{d['tl']} IS {ric}")

                for r in grp['renforts']:
                    if not any(r['id'] in [x['id'] for x in grp['principaux']] for _ in [0]):
                        css = 'rok' if r['nb'] >= r['tl'] else 'rko'
                        ic  = '✅' if r['nb'] >= r['tl'] else '⏳'
                        html_parts.append(
                            f"<tr class='{css}'><td>{r['nom']}</td><td>{r['nb']}</td>"
                            f"<td>{r['tl']}</td><td>{max(0,r['tl']-r['nb'])}</td><td>{ic}</td></tr>")

                html_parts.append("</table>")
                wa_parts.append("")

            wa_parts.append("")

        # Bilan
        html_parts.append(
            f"<div class='bilan'><b>📈 Bilan global</b><br>"
            f"DPS principaux : <b>{len(principaux)}</b> &nbsp;|&nbsp; "
            f"Incomplets : <b>{nb_incomplets}</b> &nbsp;|&nbsp; "
            f"Renforts créés : <b>{len(renforts)}</b><br>"
            f"IS requis : <b>{total_tl}</b> &nbsp;|&nbsp; "
            f"Engagés : <b>{total_eng}</b> &nbsp;|&nbsp; "
            f"Manquants : <b style='color:{'red' if total_man else 'green'}'>{total_man}</b></div>")

        wa_parts += [
            "─────────────────────",
            f"📈 *BILAN SEMAINE {self.current_semaine}*",
            f"DPS principaux : {len(principaux)} | Incomplets : {nb_incomplets}",
            f"Renforts créés : {len(renforts)}",
            f"IS requis : {total_tl} | Engagés : {total_eng} | Manquants : {total_man}",
        ]

        html_final = "\n".join(html_parts)
        wa_final   = "\n".join(wa_parts)

        # ── Dialogue ─────────────────────────────────────────────────
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Synthèse — Semaine {self.current_semaine}")
        dlg.resize(750, 580)
        layout = QVBoxLayout(dlg)

        browser = QTextBrowser()
        browser.setHtml(html_final)
        layout.addWidget(browser)

        btns = QHBoxLayout()

        btn_wa = QPushButton("📱 Copier pour WhatsApp")
        btn_wa.setStyleSheet("padding:6px 14px; background:#25D366; color:white; font-weight:bold;")
        def copier_wa():
            QApplication.clipboard().setText(wa_final)
            btn_wa.setText("✅ Copié !")
        btn_wa.clicked.connect(copier_wa)
        btns.addWidget(btn_wa)

        btn_print = QPushButton("🖨️ Imprimer")
        btn_print.setStyleSheet("padding:6px 14px; background:#1E3C72; color:white; font-weight:bold;")
        def imprimer():
            try:
                from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
                printer = QPrinter(QPrinter.PrinterMode.HighResolution)
                pd = QPrintDialog(printer, dlg)
                if pd.exec():
                    browser.print(printer)
            except Exception as ex:
                QMessageBox.warning(dlg, "Impression", f"Impression impossible :\n{ex}")
        btn_print.clicked.connect(imprimer)
        btns.addWidget(btn_print)

        btns.addStretch()
        btn_close = QPushButton("Fermer")
        btn_close.clicked.connect(dlg.accept)
        btns.addWidget(btn_close)

        layout.addLayout(btns)
        dlg.exec()

    def ouvrir_apropos(self):
        import os
        from PyQt6.QtGui import QPixmap
        dlg = QDialog(self)
        dlg.setWindowTitle("À propos — PHARE")
        dlg.setFixedSize(500, 240)
        main_layout = QVBoxLayout(dlg)

        body = QHBoxLayout()
        body.setSpacing(18)
        body.setContentsMargins(10, 10, 10, 0)

        # Logo gauche
        logo_path = resource_path(os.path.join("pictures", "PHARE Logo.png"))
        if os.path.exists(logo_path):
            lbl_logo = QLabel()
            pixmap = QPixmap(logo_path).scaledToWidth(145, Qt.TransformationMode.SmoothTransformation)
            lbl_logo.setPixmap(pixmap)
            lbl_logo.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)
            lbl_logo.setFixedWidth(150)
            body.addWidget(lbl_logo)

        # Séparateur vertical
        sep = QWidget()
        sep.setFixedWidth(1)
        sep.setStyleSheet("background:#ddd;")
        body.addWidget(sep)

        # Infos droite
        right = QVBoxLayout()
        right.setSpacing(4)

        lbl_app = QLabel(f"<b style='font-size:14px;color:#1E3C72'>{APP_NAME}</b><br>"
                         f"<span style='font-size:10px;color:#888'>"
                         f"Plateforme Harmonisée d'Affectation<br>des Renforts et des Effectifs</span>")
        lbl_ver = QLabel(f"Version <b>{VERSION}</b> &nbsp;·&nbsp; {BUILD}")
        lbl_ver.setStyleSheet("color:#555; font-size:11px;")

        sep2 = QWidget(); sep2.setFixedHeight(1); sep2.setStyleSheet("background:#ddd;")

        lbl_dev = QLabel(
            "<b>Développé par</b><br>"
            "Vachon Marc-Olivier<br>"
            "<br>"
            "© 2026 Hawkus Corp<br>"
            "<i>Tous droits réservés</i><br>"
            "<br>"
            "<b>Mis à disposition à titre d'essai</b><br>"
            "pour l'Antenne de Lannion<br>"
            "<span style='color:#888'>(22 LNP — Côtes-d'Armor)</span>"
        )
        lbl_dev.setWordWrap(True)
        lbl_dev.setStyleSheet("font-size: 11px;")

        sep3 = QWidget(); sep3.setFixedHeight(1); sep3.setStyleSheet("background:#ddd;")
        lbl_tech = QLabel("<span style='color:#aaa;font-size:10px'>Python 3 · PyQt6 · PostgreSQL (Supabase) · pandas</span>")

        right.addWidget(lbl_app)
        right.addWidget(lbl_ver)
        right.addWidget(sep2)
        right.addWidget(lbl_dev)
        right.addStretch()
        right.addWidget(sep3)
        right.addWidget(lbl_tech)

        body.addLayout(right)
        main_layout.addLayout(body)

        btn = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        btn.rejected.connect(dlg.reject)
        main_layout.addWidget(btn)
        dlg.exec()
